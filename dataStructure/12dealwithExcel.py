import openpyxl

def openExcel():
    wb=openpyxl.load_workbook('example.xlsx')
    print(type(wb))
#openExcel()

def doSomeThingAboutSheet():
    wb=openpyxl.load_workbook('example.xlsx')
    print(wb.get_sheet_names())
    sheet=wb.get_sheet_by_name('Sheet3')
    print(sheet)
    print(type(sheet))
    print(sheet.title)
    anotherSheet=wb.get_active_sheet()
    print(anotherSheet)
#doSomeThingAboutSheet()
def getTheValueOfCell():
    wb=openpyxl.load_workbook('example.xlsx')
    sheet=wb.get_sheet_by_name('Sheet1')
    print(sheet['A1'])
    print(sheet['A1'].value)
    c=sheet['C1']
    print(c)
    print(c.coordinate)#coordinate this method can get the location of Excel

#getTheValueOfCell()

def getTheSheetSize():
    wb=openpyxl.load_workbook('example.xlsx')
    sheet=wb.get_sheet_by_name('Sheet1')
    print(str(sheet.get_highest_row()))
    print(str(sheet.get_highest_column()))
#getTheSheetSize()